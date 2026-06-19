"""回写结果到 Excel: 结果列写 ok/错误, 失败行整行染红, 成功行清色。
对应修正版 VBA 的 MarkRow + 结果列逻辑。"""
from __future__ import annotations

import logging
import os

import openpyxl
from openpyxl.styles import PatternFill

from ..core.models import PlanRow

log = logging.getLogger(__name__)

# 与 VBA RGB(255,199,206) 一致的浅红("差"样式)
FAIL_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
NO_FILL = PatternFill(fill_type=None)


class ExcelWriter:
    def __init__(self, cfg: dict):
        self.cfg = cfg
        self.xcfg = cfg["excel"]

    def write(self, rows: list[PlanRow]) -> str:
        src = self.xcfg["path"]
        wb = openpyxl.load_workbook(src)   # 保留格式(不加 data_only)
        ws = wb[self.xcfg["sheet"]]

        result_col = self.xcfg["result_col"]
        last_col = result_col              # 染色范围 A..结果列
        # master 中被 LN 带出的字段也写回(原 get.data 行为)
        ln_master = {
            spec["col"]: name
            for name, spec in self.cfg["columns"]["master"].items()
            if spec.get("source") == "ln"
        }

        for row in rows:
            r = row.row_index
            # 物料规范化回写
            ws.cell(row=r, column=self.cfg["columns"]["key"]["item"]["col"], value=row.item)
            # 带出的主数据回写
            for col, name in ln_master.items():
                if name in row.master:
                    ws.cell(row=r, column=col, value=row.master[name])
            # 结果列
            ws.cell(row=r, column=result_col, value="ok" if row.ok else row.message)
            # 整行染色
            fill = NO_FILL if row.ok else FAIL_FILL
            for c in range(1, last_col + 1):
                ws.cell(row=r, column=c).fill = fill

        out = self._result_path(src)
        wb.save(out)
        log.info("结果已写入: %s", out)
        return out

    def _result_path(self, src: str) -> str:
        cfg_out = self.cfg.get("output", {}).get("result_path")
        if cfg_out:
            return cfg_out
        base, ext = os.path.splitext(src)
        return f"{base}_result{ext}"
