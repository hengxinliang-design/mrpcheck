"""读 Excel -> PlanRow 列表。按 config 的列号映射, 并校验表头是否漂移。"""
from __future__ import annotations

import logging

import openpyxl
from openpyxl.utils import get_column_letter

from ..core.models import PlanRow

log = logging.getLogger(__name__)


class ExcelReader:
    def __init__(self, cfg: dict):
        self.cfg = cfg
        self.xcfg = cfg["excel"]
        self.cols = cfg["columns"]

    def read(self) -> tuple[str, list[PlanRow]]:
        path = self.xcfg["path"]
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb[self.xcfg["sheet"]]

        method = self._cell_str(ws[self.xcfg["method_cell"]].value)
        self._validate_headers(ws)

        rows: list[PlanRow] = []
        first = self.xcfg["first_data_row"]
        for r in range(first, ws.max_row + 1):
            row = self._read_row(ws, r)
            if row.is_blank():
                continue
            rows.append(row)

        log.info("读取 %s!%s: 方案=%s, 有效行=%d", path, self.xcfg["sheet"], method, len(rows))
        return method, rows

    # ------------------------------------------------------------------
    def _read_row(self, ws, r: int) -> PlanRow:
        key = self.cols["key"]
        row = PlanRow(
            row_index=r,
            factory=self._at(ws, r, key["factory"]["col"]),
            item=self._at(ws, r, key["item"]["col"]),
            customer=self._at(ws, r, key["customer"]["col"]),
            customer_item=self._at(ws, r, key["customer_item"]["col"]),
        )
        for name, spec in self.cols["master"].items():
            row.master[name] = self._at(ws, r, spec["col"])

        pcfg = self.cols["periods"]
        start = pcfg["first_col"]
        row.periods = [self._at(ws, r, start + i) for i in range(pcfg["count"])]
        return row

    def _validate_headers(self, ws) -> None:
        """表头与 config 不一致时告警(不阻断), 替代原 VBA 的硬编码列号。"""
        hr = self.xcfg["header_row"]
        specs = list(self.cols["key"].values()) + list(self.cols["master"].values())
        for spec in specs:
            actual = self._cell_str(ws.cell(row=hr, column=spec["col"]).value)
            expected = spec.get("header", "")
            if expected and actual != expected:
                log.warning(
                    "表头漂移: 列%s 期望「%s」实际「%s」(列号映射可能需更新 config)",
                    get_column_letter(spec["col"]), expected, actual,
                )

    def _at(self, ws, r: int, c: int) -> str:
        return self._cell_str(ws.cell(row=r, column=c).value)

    @staticmethod
    def _cell_str(v) -> str:
        if v is None:
            return ""
        if isinstance(v, float) and v.is_integer():
            return str(int(v))
        return str(v).strip()
