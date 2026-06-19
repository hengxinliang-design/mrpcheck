import logging

import openpyxl

from mps_ln_import.adapters.excel_reader import ExcelReader
from mps_ln_import.adapters.excel_writer import FAIL_FILL, ExcelWriter
from .conftest import build_workbook, make_plan_row


def test_reader_maps_columns(cfg):
    build_workbook(cfg, [
        {"factory": "PG3", "item": "A1", "customer": "Q1", "customer_item": "K1",
         "master": {"sfst": "100"}, "periods": [str(i) for i in range(21)]},
        {},  # 空行应被跳过
        {"factory": "PG4", "item": "A2", "customer": "Q2", "customer_item": "K2",
         "periods": ["0"] * 21},
    ], method="ACT")

    method, rows = ExcelReader(cfg).read()
    assert method == "ACT"
    assert len(rows) == 2                      # 空行跳过
    assert rows[0].factory == "PG3"
    assert rows[0].master["sfst"] == "100"
    assert rows[0].periods[20] == "20"
    assert rows[0].row_index == cfg["excel"]["first_data_row"]
    assert rows[1].item == "A2"


def test_reader_header_drift_warns(cfg, caplog):
    build_workbook(cfg, [{"factory": "PG3", "item": "A1", "periods": ["0"] * 21}],
                   drift_col=cfg["columns"]["key"]["item"]["col"])
    with caplog.at_level(logging.WARNING):
        ExcelReader(cfg).read()
    assert any("表头漂移" in r.message for r in caplog.records)


def test_writer_marks_fail_red_and_ok_clear(cfg):
    build_workbook(cfg, [
        {"factory": "PG3", "item": "A1", "customer": "Q1", "customer_item": "K1", "periods": ["0"] * 21},
        {"factory": "PG4", "item": "A2", "customer": "Q2", "customer_item": "K2", "periods": ["0"] * 21},
    ])
    ok = make_plan_row(5, item="A1NORM"); ok.ok = True; ok.message = "ok"
    bad = make_plan_row(6); bad.ok = False; bad.message = "import失败: X"

    out = ExcelWriter(cfg).write([ok, bad])

    wb = openpyxl.load_workbook(out)
    ws = wb[cfg["excel"]["sheet"]]
    rc = cfg["excel"]["result_col"]
    item_col = cfg["columns"]["key"]["item"]["col"]

    assert ws.cell(row=5, column=rc).value == "ok"
    assert ws.cell(row=5, column=item_col).value == "A1NORM"        # 规范化回写
    assert ws.cell(row=5, column=1).fill.fill_type is None          # 成功清色

    assert "import失败" in ws.cell(row=6, column=rc).value
    assert ws.cell(row=6, column=1).fill.start_color.rgb == FAIL_FILL.start_color.rgb  # 失败染红


def test_writer_result_path_default(cfg):
    cfg["output"]["result_path"] = ""
    build_workbook(cfg, [{"factory": "PG3", "item": "A1", "periods": ["0"] * 21}])
    row = make_plan_row(5); row.ok = True
    out = ExcelWriter(cfg).write([row])
    assert out.endswith("_result.xlsx")
