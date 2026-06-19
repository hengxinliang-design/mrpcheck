"""测试夹具: 加载真实 config(保证列映射与生产一致), 并按需合成 Excel。"""
from __future__ import annotations

import copy
import os

import openpyxl
import pytest
import yaml

from mps_ln_import.core.models import PlanRow

CONFIG_PATH = os.path.join(os.path.dirname(__file__), "..", "config.yaml")


@pytest.fixture
def base_cfg() -> dict:
    """真实 config.yaml 的深拷贝, 测试里随意改不影响文件。"""
    with open(CONFIG_PATH, encoding="utf-8") as f:
        return copy.deepcopy(yaml.safe_load(f))


@pytest.fixture
def cfg(base_cfg, tmp_path) -> dict:
    """指向临时路径的 config, 避免碰到桌面真实文件。"""
    c = base_cfg
    c["excel"]["path"] = str(tmp_path / "src.xlsx")
    c["output"]["result_path"] = str(tmp_path / "src_result.xlsx")
    c["output"]["log_path"] = str(tmp_path / "run.log")
    return c


def make_plan_row(row_index: int, factory="PG3", item="a04-x", customer="Q1",
                  customer_item="K1", periods=None) -> PlanRow:
    periods = periods if periods is not None else ["0"] * 21
    return PlanRow(
        row_index=row_index, factory=factory, item=item,
        customer=customer, customer_item=customer_item,
        master={}, periods=list(periods),
    )


def build_workbook(cfg: dict, rows: list[dict], method="ACT", drift_col: int | None = None) -> str:
    """按 cfg 列映射合成一个工作簿并保存到 cfg.excel.path。

    rows: 每个元素 {factory,item,customer,customer_item, master:{}, periods:[...]}。
    drift_col: 若指定, 把该列表头写成错误值, 用于测试表头漂移告警。
    """
    xcfg = cfg["excel"]
    cols = cfg["columns"]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = xcfg["sheet"]
    ws[xcfg["method_cell"]] = method

    # 表头(取自 config 的 header, 保证默认不漂移)
    hr = xcfg["header_row"]
    for spec in list(cols["key"].values()) + list(cols["master"].values()):
        ws.cell(row=hr, column=spec["col"], value=spec.get("header", ""))
    if drift_col is not None:
        ws.cell(row=hr, column=drift_col, value="表头被改了")

    # 数据行
    pcfg = cols["periods"]
    for i, rd in enumerate(rows):
        r = xcfg["first_data_row"] + i
        ws.cell(row=r, column=cols["key"]["factory"]["col"], value=rd.get("factory", ""))
        ws.cell(row=r, column=cols["key"]["item"]["col"], value=rd.get("item", ""))
        ws.cell(row=r, column=cols["key"]["customer"]["col"], value=rd.get("customer", ""))
        ws.cell(row=r, column=cols["key"]["customer_item"]["col"], value=rd.get("customer_item", ""))
        for name, val in rd.get("master", {}).items():
            ws.cell(row=r, column=cols["master"][name]["col"], value=val)
        for j, p in enumerate(rd.get("periods", [])):
            ws.cell(row=r, column=pcfg["first_col"] + j, value=p)

    path = xcfg["path"]
    wb.save(path)
    return path
